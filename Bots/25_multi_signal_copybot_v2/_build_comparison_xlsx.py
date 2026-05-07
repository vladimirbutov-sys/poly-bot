"""
Build portfolio comparison Excel: Our bot vs Denizz.
Outputs: 2026-04-16_portfolio_comparison.xlsx
"""
import sys
import os
import json
import time
import requests
from datetime import datetime
from collections import defaultdict

sys.path.insert(0, "c:/Users/Honor/Desktop/Polymarket/Bots/25_multi_signal_copybot_v2")
import filters

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
except ImportError:
    os.system("pip install openpyxl")
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

DENIZZ_WALLET = "0xbaa2bcb5439e985ce4ccf815b4700027d1b92c73"
DATA_API = "https://data-api.polymarket.com"
BOT_DIR = "c:/Users/Honor/Desktop/Polymarket/Bots/25_multi_signal_copybot_v2"
OUTPUT = os.path.join(BOT_DIR, "2026-04-16_portfolio_comparison.xlsx")

# ── Load positions ──────────────────────────────────────────────────────
with open(os.path.join(BOT_DIR, "positions.json")) as f:
    raw = json.load(f)

positions = []
our_cids = set()
our_tids = set()
for key, p in raw["positions"].items():
    if p.get("status") == "merged_into_primary":
        continue
    if p.get("cost_usd", 0) == 0:
        continue
    positions.append(p)
    our_cids.add(p["condition_id"])
    our_tids.add(p["token_id"])

print(f"Loaded {len(positions)} positions (excl merged/zero)")
print(f"Unique condition_ids: {len(our_cids)}, token_ids: {len(our_tids)}")

# ── Fetch ALL Denizz trades via pagination ──────────────────────────────
print("\nFetching Denizz trade history...")
all_denizz_trades = []
offset = 0
batch_size = 500

while offset < 50000:
    url = f"{DATA_API}/activity?user={DENIZZ_WALLET}&limit={batch_size}&offset={offset}"
    try:
        r = requests.get(url, timeout=30)
        if r.status_code != 200:
            print(f"  API returned {r.status_code} at offset {offset}, stopping pagination")
            break
        data = r.json()
        if not data:
            break

        matches = 0
        for d in data:
            if d.get("conditionId") in our_cids or d.get("asset") in our_tids:
                all_denizz_trades.append(d)
                matches += 1

        print(f"  offset={offset}: {len(data)} trades, {matches} matching ({len(all_denizz_trades)} total)")

        if len(data) < batch_size:
            break
        offset += batch_size
        time.sleep(0.3)
    except Exception as e:
        print(f"  Error at offset {offset}: {e}")
        break

print(f"Total Denizz trades matching our markets: {len(all_denizz_trades)}")

# ── Aggregate Denizz trades by (conditionId, asset) ────────────────────
denizz_agg = {}  # key=(cid, tid) -> {buys: [(price, size)], sells: [(price, size)]}

for t in all_denizz_trades:
    cid = t.get("conditionId", "")
    tid = t.get("asset", "")
    key = (cid, tid)

    if key not in denizz_agg:
        denizz_agg[key] = {"buys": [], "sells": []}

    side = t.get("side", "").upper()
    price = float(t.get("price", 0) or 0)
    size = float(t.get("size", 0) or 0)

    if side == "BUY" and size > 0:
        denizz_agg[key]["buys"].append((price, size))
    elif side == "SELL" and size > 0:
        denizz_agg[key]["sells"].append((price, size))

# Compute summary stats per (cid, tid)
denizz_data = {}
for key, trades in denizz_agg.items():
    buys = trades["buys"]
    sells = trades["sells"]

    total_buy_shares = sum(s for _, s in buys)
    total_buy_cost = sum(p * s for p, s in buys)
    total_sell_shares = sum(s for _, s in sells)
    total_sell_revenue = sum(p * s for p, s in sells)

    avg_buy = total_buy_cost / total_buy_shares if total_buy_shares > 0 else 0
    avg_sell = total_sell_revenue / total_sell_shares if total_sell_shares > 0 else 0
    remaining = max(total_buy_shares - total_sell_shares, 0)

    denizz_data[key] = {
        "avg_buy": avg_buy,
        "avg_sell": avg_sell,
        "total_bought_usd": total_buy_cost,
        "total_sold_usd": total_sell_revenue,
        "total_buy_shares": total_buy_shares,
        "total_sell_shares": total_sell_shares,
        "remaining_shares": remaining,
    }

print(f"Denizz data aggregated for {len(denizz_data)} (cid, tid) pairs")

# ── Fetch current prices for open positions ─────────────────────────────
current_prices = {}
open_tids = set()
for p in positions:
    if p["status"] == "open":
        open_tids.add(p["token_id"])

print(f"\nFetching current prices for {len(open_tids)} open token_ids...")
for i, tid in enumerate(open_tids):
    print(f"  [{i+1}/{len(open_tids)}] Orderbook for tid={tid[:20]}...")
    try:
        prices = filters.get_orderbook_prices(tid)
        current_prices[tid] = prices[0] if prices else 0
    except Exception as e:
        print(f"    Error: {e}")
        current_prices[tid] = 0
    time.sleep(0.3)

# ── Build rows ──────────────────────────────────────────────────────────
rows = []
for p in positions:
    cid = p["condition_id"]
    tid = p["token_id"]

    # Our data
    sells = p.get("sells", [])
    valid_sells = [s for s in sells if s.get("shares", 0) > 0]

    total_sell_shares_us = sum(s.get("shares", 0) for s in valid_sells)
    total_sell_revenue_us = sum(s.get("revenue", 0) for s in valid_sells)
    realized_pnl_us = sum(s.get("pnl", 0) for s in valid_sells)

    # Weighted avg exit
    if total_sell_shares_us > 0:
        avg_exit_us = sum(s.get("price", 0) * s.get("shares", 0) for s in valid_sells) / total_sell_shares_us
    else:
        avg_exit_us = 0

    cur_price = current_prices.get(tid, 0)

    # Remaining shares
    initial_shares = p.get("cost_usd", 0) / p.get("avg_entry", 1) if p.get("avg_entry", 0) > 0 else 0
    remaining_shares_us = p.get("size_shares", 0)

    # Unrealized PnL
    if p["status"] == "open" and remaining_shares_us > 0 and cur_price > 0:
        unrealized_pnl_us = (cur_price - p.get("avg_entry", 0)) * remaining_shares_us
    else:
        unrealized_pnl_us = 0

    # Total PnL
    if p["status"] != "open" and "final_pnl" in p:
        total_pnl_us = p["final_pnl"]
    else:
        total_pnl_us = realized_pnl_us + unrealized_pnl_us

    cost_usd = p.get("cost_usd", 0)
    pnl_pct_us = (total_pnl_us / cost_usd * 100) if cost_usd > 0 else 0

    # Denizz data
    dd = denizz_data.get((cid, tid))
    if dd and dd["total_buy_shares"] > 0:
        denizz_avg_entry = dd["avg_buy"]
        denizz_invested = dd["total_bought_usd"]
        denizz_avg_exit = dd["avg_sell"]
        denizz_remaining_shares = dd["remaining_shares"]
        denizz_remaining_usd = denizz_remaining_shares * cur_price if cur_price > 0 else denizz_remaining_shares * denizz_avg_entry

        # Denizz realized PnL
        if dd["total_sell_shares"] > 0:
            cost_of_sold = (dd["total_sell_shares"] / dd["total_buy_shares"]) * dd["total_bought_usd"]
            denizz_realized_pnl = dd["total_sold_usd"] - cost_of_sold
        else:
            denizz_realized_pnl = 0

        # Denizz unrealized
        if denizz_remaining_shares > 0 and cur_price > 0:
            denizz_unrealized = (cur_price - denizz_avg_entry) * denizz_remaining_shares
        else:
            denizz_unrealized = 0

        denizz_total_pnl = denizz_realized_pnl + denizz_unrealized
        denizz_pnl_pct = (denizz_total_pnl / denizz_invested * 100) if denizz_invested > 0 else 0
    else:
        denizz_avg_entry = 0
        denizz_invested = 0
        denizz_avg_exit = 0
        denizz_remaining_usd = 0
        denizz_realized_pnl = 0
        denizz_pnl_pct = 0

    # Date opened
    ts = p.get("timestamp", "")
    try:
        dt = datetime.fromisoformat(ts.replace("+00:00", "")).strftime("%Y-%m-%d %H:%M")
    except:
        dt = ts[:16] if ts else "N/A"

    row = {
        "Market": p.get("title", ""),
        "Outcome": p.get("outcome", ""),
        "Status": p.get("status", ""),
        "Signal Player": p.get("signal_player", "") or "manual",
        "Date Opened": dt,
        "Our Avg Entry": round(p.get("avg_entry", 0), 4),
        "Our Cost USD": round(cost_usd, 2),
        "Our Shares": round(remaining_shares_us if p["status"] == "open" else initial_shares, 2),
        "Our Avg Exit": round(avg_exit_us, 4),
        "Our Revenue": round(total_sell_revenue_us, 2),
        "Our Realized PnL": round(realized_pnl_us, 2),
        "Our Unrealized PnL": round(unrealized_pnl_us, 2),
        "Our Total PnL": round(total_pnl_us, 2),
        "Our PnL %": round(pnl_pct_us, 2),
        "Denizz Avg Entry": round(denizz_avg_entry, 4),
        "Denizz Total Invested": round(denizz_invested, 2),
        "Denizz Avg Exit": round(denizz_avg_exit, 4),
        "Denizz Remaining USD": round(denizz_remaining_usd, 2),
        "Denizz Realized PnL": round(denizz_realized_pnl, 2),
        "Denizz PnL %": round(denizz_pnl_pct, 2),
        "Current Price": round(cur_price, 4),
        "Event Slug": p.get("event_slug", ""),
    }
    rows.append(row)

# Sort: open first, then by date
rows.sort(key=lambda r: (0 if r["Status"] == "open" else 1, r["Date Opened"]))

print(f"\nBuilt {len(rows)} rows. Creating Excel...")

# ── Create Excel ────────────────────────────────────────────────────────
wb = Workbook()

# ── Sheet 1: All Positions ──────────────────────────────────────────────
ws1 = wb.active
ws1.title = "All Positions"

headers = list(rows[0].keys())
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font_white = Font(bold=True, size=11, color="FFFFFF")

green_font = Font(color="006100")
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_font = Font(color="9C0006")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

# Write headers
for col_idx, h in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col_idx, value=h)
    cell.font = header_font_white
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

pnl_cols = {"Our Realized PnL", "Our Unrealized PnL", "Our Total PnL", "Our PnL %",
            "Denizz Realized PnL", "Denizz PnL %"}
money_cols = {"Our Cost USD", "Our Revenue", "Our Realized PnL", "Our Unrealized PnL",
              "Our Total PnL", "Denizz Total Invested", "Denizz Remaining USD", "Denizz Realized PnL"}
pct_cols = {"Our PnL %", "Denizz PnL %"}
price_cols = {"Our Avg Entry", "Our Avg Exit", "Denizz Avg Entry", "Denizz Avg Exit", "Current Price"}

for row_idx, row_data in enumerate(rows, 2):
    for col_idx, h in enumerate(headers, 1):
        val = row_data[h]
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)

        if h in money_cols:
            cell.number_format = '#,##0.00'
        elif h in pct_cols:
            cell.number_format = '#,##0.00'
        elif h in price_cols:
            cell.number_format = '0.0000'

        if h in pnl_cols and isinstance(val, (int, float)):
            if val > 0:
                cell.font = green_font
                cell.fill = green_fill
            elif val < 0:
                cell.font = red_font
                cell.fill = red_fill

        if h == "Status":
            if val == "open":
                cell.fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
            elif val == "won":
                cell.fill = green_fill
            elif val == "lost":
                cell.fill = red_fill

# Auto-width
for col_idx, h in enumerate(headers, 1):
    max_len = len(str(h))
    for row_idx in range(2, min(len(rows) + 2, 50)):  # sample first 50 rows
        cell_val = ws1.cell(row=row_idx, column=col_idx).value
        if cell_val is not None:
            max_len = max(max_len, len(str(cell_val)))
    ws1.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 45)

ws1.freeze_panes = "A2"

# ── Sheet 2: Summary ───────────────────────────────────────────────────
ws2 = wb.create_sheet("Summary")

total_pos = len(rows)
open_pos = sum(1 for r in rows if r["Status"] == "open")
closed_pos = total_pos - open_pos
sold_pos = sum(1 for r in rows if r["Status"] == "sold")
won_pos = sum(1 for r in rows if r["Status"] == "won")
lost_pos = sum(1 for r in rows if r["Status"] == "lost")

total_cost = sum(r["Our Cost USD"] for r in rows)
total_realized = sum(r["Our Realized PnL"] for r in rows)
total_unrealized = sum(r["Our Unrealized PnL"] for r in rows)
total_pnl_all = sum(r["Our Total PnL"] for r in rows)
total_revenue = sum(r["Our Revenue"] for r in rows)

closed_rows = [r for r in rows if r["Status"] != "open"]
wins = sum(1 for r in closed_rows if r["Our Total PnL"] > 0)
win_rate = (wins / len(closed_rows) * 100) if closed_rows else 0

avg_pnl_pct_us = sum(r["Our PnL %"] for r in rows) / len(rows) if rows else 0
denizz_rows_with_data = [r for r in rows if r["Denizz Total Invested"] > 0]
avg_pnl_pct_denizz = sum(r["Denizz PnL %"] for r in denizz_rows_with_data) / len(denizz_rows_with_data) if denizz_rows_with_data else 0

total_denizz_invested = sum(r["Denizz Total Invested"] for r in rows)
total_denizz_realized = sum(r["Denizz Realized PnL"] for r in rows)

summary_data = [
    ("СВОДКА ПОРТФЕЛЯ", ""),
    ("Дата отчёта", "2026-04-16"),
    ("", ""),
    ("ПОЗИЦИИ", ""),
    ("Всего позиций", total_pos),
    ("Открытых", open_pos),
    ("Закрытых (sold)", sold_pos),
    ("Выиграно (won)", won_pos),
    ("Проиграно (lost)", lost_pos),
    ("", ""),
    ("НАШИ РЕЗУЛЬТАТЫ", ""),
    ("Общие вложения", round(total_cost, 2)),
    ("Общая выручка от продаж", round(total_revenue, 2)),
    ("Реализованный PnL", round(total_realized, 2)),
    ("Нереализованный PnL (open)", round(total_unrealized, 2)),
    ("Общий PnL", round(total_pnl_all, 2)),
    ("ROI %", round(total_pnl_all / total_cost * 100, 2) if total_cost else 0),
    ("Win Rate (закрытые)", f"{round(win_rate, 1)}%"),
    ("Средний PnL% на позицию", round(avg_pnl_pct_us, 2)),
    ("", ""),
    ("DENIZZ (на тех же рынках)", ""),
    ("Позиций с данными Denizz", len(denizz_rows_with_data)),
    ("Общие вложения Denizz", round(total_denizz_invested, 2)),
    ("Реализованный PnL Denizz", round(total_denizz_realized, 2)),
    ("Средний PnL% Denizz", round(avg_pnl_pct_denizz, 2)),
    ("", ""),
    ("СРАВНЕНИЕ", ""),
    ("Наш средний PnL%", round(avg_pnl_pct_us, 2)),
    ("Denizz средний PnL%", round(avg_pnl_pct_denizz, 2)),
    ("Разница (мы - Denizz)", round(avg_pnl_pct_us - avg_pnl_pct_denizz, 2)),
    ("", ""),
    ("ПРИМЕЧАНИЕ", ""),
    ("", "Данные Denizz получены из API (последние ~3500 сделок)."),
    ("", "Для закрытых позиций Denizz данные могут быть неполными."),
    ("", "Позиции без данных Denizz показаны с нулями."),
]

for row_idx, (label, val) in enumerate(summary_data, 1):
    cell_a = ws2.cell(row=row_idx, column=1, value=label)
    cell_b = ws2.cell(row=row_idx, column=2, value=val)

    if label in ("СВОДКА ПОРТФЕЛЯ", "ПОЗИЦИИ", "НАШИ РЕЗУЛЬТАТЫ", "DENIZZ (на тех же рынках)", "СРАВНЕНИЕ", "ПРИМЕЧАНИЕ"):
        cell_a.font = Font(bold=True, size=13)
        cell_a.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        cell_b.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    else:
        cell_a.font = Font(size=11)

    if isinstance(val, (int, float)):
        if "PnL" in label or "ROI" in label:
            if val > 0:
                cell_b.font = green_font
                cell_b.fill = green_fill
            elif val < 0:
                cell_b.font = red_font
                cell_b.fill = red_fill
        cell_b.number_format = '#,##0.00'

ws2.column_dimensions["A"].width = 35
ws2.column_dimensions["B"].width = 50

# ── Save ────────────────────────────────────────────────────────────────
wb.save(OUTPUT)
print(f"\n{'='*60}")
print(f"Saved to: {OUTPUT}")
print(f"  {total_pos} positions, {open_pos} open, {closed_pos} closed")
print(f"  Our total PnL: ${total_pnl_all:.2f}")
print(f"  Win rate: {win_rate:.1f}%")
print(f"  Denizz data available for {len(denizz_rows_with_data)} positions")
print(f"{'='*60}")
