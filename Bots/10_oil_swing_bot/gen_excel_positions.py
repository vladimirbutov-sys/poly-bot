"""Generate Excel file with current positions and live market values."""
import os
import requests
from datetime import datetime, timezone
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers

from config import NO_TOKEN_ID, YES_TOKEN_ID, NO_ENTRY_STEPS
import tracker

BOT_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH = os.path.join(BOT_DIR, "Oil_Bot_Positions.xlsx")


def get_live_prices():
    """Fetch current bid/ask for YES and NO."""
    prices = {}
    for name, token_id in [("YES", YES_TOKEN_ID), ("NO", NO_TOKEN_ID)]:
        try:
            resp = requests.get(
                "https://clob.polymarket.com/book",
                params={"token_id": token_id},
                timeout=10,
            )
            if resp.status_code == 200:
                book = resp.json()
                bids = book.get("bids", [])
                asks = book.get("asks", [])
                best_bid = max(float(b["price"]) for b in bids) if bids else 0.0
                best_ask = min(float(a["price"]) for a in asks) if asks else 1.0
                prices[name] = {"bid": best_bid, "ask": best_ask}
        except Exception as e:
            print(f"Error fetching {name} prices: {e}")
            prices[name] = {"bid": 0, "ask": 0}
    return prices


def get_no_profit_target(entry_step):
    if 0 <= entry_step < len(NO_ENTRY_STEPS):
        return NO_ENTRY_STEPS[entry_step][2]
    return 0.50


def generate():
    data = tracker.load()
    prices = get_live_prices()
    now = datetime.now(timezone.utc)

    wb = Workbook()

    # ===== Sheet 1: Open Positions =====
    ws = wb.active
    ws.title = "Open Positions"

    # Styles
    header_font = Font(bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    green_font = Font(color="008000", bold=True)
    red_font = Font(color="FF0000", bold=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    # Title
    ws.merge_cells('A1:J1')
    ws['A1'] = f"Oil Swing Bot — Positions (updated {now.strftime('%d.%m.%Y %H:%M UTC')})"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')

    # Market prices row
    ws.merge_cells('A2:J2')
    yes_bid = prices.get("YES", {}).get("bid", 0)
    yes_ask = prices.get("YES", {}).get("ask", 0)
    no_bid = prices.get("NO", {}).get("bid", 0)
    no_ask = prices.get("NO", {}).get("ask", 0)
    ws['A2'] = (f"YES bid/ask: {yes_bid:.3f}/{yes_ask:.3f}  |  "
                f"NO bid/ask: {no_bid:.3f}/{no_ask:.3f}")
    ws['A2'].font = Font(size=10, italic=True)
    ws['A2'].alignment = Alignment(horizontal='center')

    # Headers
    headers = [
        "Side", "Entry Price", "Current Bid", "Shares",
        "Cost ($)", "Market Value ($)", "P&L ($)", "P&L (%)",
        "Entry Step", "Profit Target",
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    # Data
    row_idx = 5
    total_cost = 0
    total_value = 0
    total_pnl = 0

    open_positions = []
    for key, pos in data["positions"].items():
        if pos.get("status") == "open":
            open_positions.append((key, pos))

    # Sort: YES first, then NO
    open_positions.sort(key=lambda x: (0 if x[1]["side"] == "YES" else 1, x[1].get("entry_step", -1)))

    for key, pos in open_positions:
        side = pos["side"]
        entry = pos.get("entry_price", 0)
        shares = pos.get("size_shares", 0)
        cost = pos.get("cost_usd", 0)
        step = pos.get("entry_step", -1)

        current_bid = prices.get(side, {}).get("bid", 0)
        market_val = shares * current_bid
        pnl = market_val - cost
        pnl_pct = (pnl / cost * 100) if cost > 0 else 0

        if side == "NO":
            target = get_no_profit_target(step)
            target_str = f"+{target:.0%}"
        else:
            target_str = "+15%"

        step_str = f"Step {step + 1}" if step >= 0 else "Legacy"

        row_data = [
            side, entry, current_bid, shares,
            cost, market_val, pnl, pnl_pct,
            step_str, target_str,
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')

            if col == 2 or col == 3:  # prices
                cell.number_format = '0.000'
            elif col == 4:  # shares
                cell.number_format = '0.00'
            elif col in (5, 6, 7):  # USD
                cell.number_format = '$#,##0.00'
            elif col == 8:  # percentage
                cell.number_format = '0.0"%"'

            # Color P&L
            if col == 7 or col == 8:
                if isinstance(val, (int, float)):
                    cell.font = green_font if val >= 0 else red_font

        total_cost += cost
        total_value += market_val
        total_pnl += pnl
        row_idx += 1

    # Totals row
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="TOTAL").font = Font(bold=True, size=11)
    for col, val in [(5, total_cost), (6, total_value), (7, total_pnl)]:
        cell = ws.cell(row=row_idx, column=col, value=val)
        cell.font = Font(bold=True, size=11)
        cell.number_format = '$#,##0.00'
        cell.border = border
        if col == 7:
            cell.font = Font(bold=True, size=11,
                             color="008000" if val >= 0 else "FF0000")

    total_pnl_pct = (total_pnl / total_cost * 100) if total_cost > 0 else 0
    cell = ws.cell(row=row_idx, column=8, value=total_pnl_pct)
    cell.number_format = '0.0"%"'
    cell.font = Font(bold=True, size=11,
                     color="008000" if total_pnl_pct >= 0 else "FF0000")
    cell.border = border

    # Balance info
    row_idx += 2
    balance = tracker.get_balance(data)
    portfolio = balance + total_value
    ws.cell(row=row_idx, column=1, value="Free Balance:").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=balance).number_format = '$#,##0.00'
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Portfolio Value:").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value=portfolio).number_format = '$#,##0.00'

    # Column widths
    widths = [8, 12, 12, 10, 12, 14, 12, 10, 12, 14]
    for i, w in enumerate(widths):
        ws.column_dimensions[chr(65 + i)].width = w

    # ===== Sheet 2: Trade History =====
    ws2 = wb.create_sheet("History")
    hist_headers = ["Side", "Entry", "Exit", "Shares", "P&L ($)", "Reason", "Closed At"]
    for col, h in enumerate(hist_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = border

    for r_idx, h in enumerate(data.get("history", []), 2):
        row_data = [
            h.get("side", ""),
            h.get("entry", 0),
            h.get("exit", 0),
            h.get("shares", 0),
            h.get("pnl", 0),
            h.get("reason", ""),
            h.get("closed_at", ""),
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=r_idx, column=col, value=val)
            cell.border = border
            if col in (2, 3):
                cell.number_format = '0.000'
            elif col == 5:
                cell.number_format = '$#,##0.00'
                if isinstance(val, (int, float)):
                    cell.font = green_font if val >= 0 else red_font

    hist_widths = [8, 10, 10, 10, 12, 20, 25]
    for i, w in enumerate(hist_widths):
        ws2.column_dimensions[chr(65 + i)].width = w

    # ===== Sheet 3: Bot Stats =====
    ws3 = wb.create_sheet("Stats")
    stats = data.get("stats", {})
    stat_rows = [
        ("Bankroll", f"${stats.get('bankroll', 0):.2f}"),
        ("Current Balance", f"${stats.get('current_balance', 0):.2f}"),
        ("Total Trades", str(stats.get('total_trades', 0))),
        ("Total P&L", f"${stats.get('total_pnl', 0):+.2f}"),
        ("Open Positions", str(len(open_positions))),
        ("Total Invested", f"${total_cost:.2f}"),
        ("Market Value", f"${total_value:.2f}"),
        ("Unrealized P&L", f"${total_pnl:+.2f}"),
        ("Portfolio Value", f"${portfolio:.2f}"),
    ]
    for r_idx, (label, val) in enumerate(stat_rows, 1):
        ws3.cell(row=r_idx, column=1, value=label).font = Font(bold=True)
        ws3.cell(row=r_idx, column=2, value=val)

    ws3.column_dimensions['A'].width = 20
    ws3.column_dimensions['B'].width = 15

    wb.save(OUTPUT_PATH)
    print(f"Excel saved: {OUTPUT_PATH}")
    print(f"Open positions: {len(open_positions)}")
    print(f"Total invested: ${total_cost:.2f}")
    print(f"Market value: ${total_value:.2f}")
    print(f"Unrealized P&L: ${total_pnl:+.2f}")


if __name__ == "__main__":
    generate()
