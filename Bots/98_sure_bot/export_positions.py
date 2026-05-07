"""Export current positions to Excel with live Polymarket status."""
import json
import time
import logging
import requests
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers

logging.basicConfig(level=logging.INFO, format="%(asctime)s %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("export")

POSITIONS_FILE = Path(__file__).parent / "positions.json"
EXCEL_FILE = Path(__file__).parent / "positions_report.xlsx"
GAMMA_API = "https://gamma-api.polymarket.com/markets"


def load_positions():
    with open(POSITIONS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)


def fetch_market_info(token_id: str) -> dict | None:
    """Fetch current market data from Gamma API using clob_token_ids."""
    try:
        resp = requests.get(GAMMA_API, params={"clob_token_ids": token_id, "limit": 1}, timeout=10)
        if resp.ok:
            data = resp.json()
            if data and len(data) > 0:
                return data[0]
    except Exception as e:
        log.warning("API error for %s: %s", token_id[:16], e)
    return None


def get_market_status(market: dict, token_id: str) -> tuple:
    """Extract current price and resolution status from market data.
    Returns (current_price, market_status, end_date, resolved_flag).
    """
    if not market:
        return None, "unknown", None, False

    # End date
    end_date = market.get("endDate", "") or market.get("end_date", "")

    # Check if resolved
    closed = market.get("closed", False)
    resolved = market.get("resolved", False)

    # Current price for our token
    current_price = None
    try:
        tokens_str = market.get("clobTokenIds", "") or ""
        prices_str = market.get("outcomePrices", "") or ""
        if tokens_str and prices_str:
            # Parse JSON arrays stored as strings
            if isinstance(tokens_str, str):
                tokens = json.loads(tokens_str)
            else:
                tokens = tokens_str
            if isinstance(prices_str, str):
                prices = json.loads(prices_str)
            else:
                prices = prices_str
            for t, p in zip(tokens, prices):
                if t == token_id:
                    current_price = float(p)
                    break
    except (json.JSONDecodeError, ValueError, TypeError):
        pass

    # Determine status
    if resolved or closed:
        # Check winner
        winner = market.get("winner", "")
        if winner:
            status = "resolved"
        else:
            status = "closed"
    else:
        uma = market.get("umaResolutionStatus", "") or ""
        if uma == "proposed":
            status = "resolving (~2h)"
        else:
            status = "active"

    return current_price, status, end_date, bool(resolved or closed)


def build_excel(positions_data: dict):
    """Build Excel workbook with all positions + live data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Positions"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    won_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    lost_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    resolving_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    selling_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # light blue
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Headers
    headers = [
        "#", "Title", "Entry Price", "Current Price", "Price Change",
        "Cost USD", "Shares", "Expected PnL", "Outcome", "Realized PnL",
        "Status (bot)", "Market Status",
        "End Date", "Neg Risk", "Bought At"
    ]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = thin_border

    # Sort positions by timestamp
    positions = list(positions_data["positions"].values())
    positions.sort(key=lambda p: p.get("timestamp", ""))

    # Fetch market data by token_id (clob_token_ids gives correct market-level data)
    token_ids = set()
    for p in positions:
        tid = p.get("token_id", "")
        if tid:
            token_ids.add(tid)

    log.info("Fetching market data for %d unique tokens...", len(token_ids))
    market_cache = {}
    for i, tid in enumerate(token_ids):
        if i > 0 and i % 10 == 0:
            log.info("  ...fetched %d/%d", i, len(token_ids))
            time.sleep(0.5)  # rate limit
        market_cache[tid] = fetch_market_info(tid)
        time.sleep(0.2)
    log.info("  ...done fetching all %d markets", len(token_ids))

    # Fill rows
    total_cost = 0
    total_expected_pnl = 0
    total_realized_pnl = 0

    for idx, pos in enumerate(positions, 1):
        row = idx + 1
        token_id = pos.get("token_id", "")
        market = market_cache.get(token_id)
        current_price, mkt_status, end_date, is_resolved = get_market_status(market, token_id)

        entry_price = pos.get("entry_price", 0)
        cost = pos.get("cost_usd", 0)
        shares = pos.get("size_shares", 0)
        bot_status = pos.get("status", "open")
        neg_risk = pos.get("neg_risk", False)
        timestamp = pos.get("timestamp", "")[:16].replace("T", " ")

        # Expected PnL (if we win)
        expected_pnl = shares - cost

        # Outcome prediction based on current price
        # current_price near 1.0 = we win, near 0.0 = we lose
        outcome = ""
        realized_pnl = None

        if bot_status == "won":
            outcome = "WIN"
            realized_pnl = pos.get("pnl", shares - cost)
        elif bot_status == "lost":
            outcome = "LOSS"
            realized_pnl = pos.get("pnl", -cost)
        elif bot_status == "selling":
            sell_price = pos.get("sell_price", 0)
            if sell_price:
                realized_pnl = (sell_price - entry_price) * shares
                outcome = "SELLING" if realized_pnl >= 0 else "SELLING (-)"
            else:
                outcome = "SELLING"
        elif bot_status == "sold":
            sell_price = pos.get("sell_price", 0)
            realized_pnl = (sell_price - entry_price) * shares if sell_price else 0
            outcome = "SOLD"
        elif mkt_status == "resolved":
            # Bot hasn't updated yet but market resolved
            if current_price is not None and current_price > 0.5:
                outcome = "WIN"
                realized_pnl = shares - cost
            elif current_price is not None:
                outcome = "LOSS"
                realized_pnl = -cost
            else:
                outcome = "RESOLVED"
        elif mkt_status == "resolving (~2h)":
            if current_price is not None and current_price > 0.5:
                outcome = "-> WIN"
                realized_pnl = shares - cost
            elif current_price is not None:
                outcome = "-> LOSS"
                realized_pnl = -cost
            else:
                outcome = "RESOLVING"

        total_cost += cost
        total_expected_pnl += expected_pnl
        if realized_pnl is not None:
            total_realized_pnl += realized_pnl

        # Parse end_date for display
        end_date_display = ""
        if end_date:
            try:
                edt = end_date.replace("Z", "+00:00")
                if "T" in edt:
                    ed = datetime.fromisoformat(edt)
                    end_date_display = ed.strftime("%Y-%m-%d %H:%M")
                else:
                    end_date_display = end_date[:10]
            except (ValueError, TypeError):
                end_date_display = str(end_date)[:16]

        # Price change (current - entry)
        price_change = None
        if current_price is not None:
            price_change = current_price - entry_price

        values = [
            idx,
            pos.get("title", ""),
            entry_price,
            current_price,
            price_change,
            cost,
            shares,
            expected_pnl,
            outcome,
            realized_pnl,
            bot_status,
            mkt_status,
            end_date_display,
            "Yes" if neg_risk else "No",
            timestamp,
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=(col == 2))

            # Row coloring
            if outcome in ("WIN", "-> WIN"):
                cell.fill = won_fill
            elif outcome in ("LOSS", "-> LOSS"):
                cell.fill = lost_fill
            elif outcome in ("SELLING", "SELLING (-)", "SOLD"):
                cell.fill = selling_fill
            elif mkt_status == "resolving (~2h)":
                cell.fill = resolving_fill

        # Format numbers
        ws.cell(row=row, column=3).number_format = '0.0000'  # entry price
        if current_price is not None:
            ws.cell(row=row, column=4).number_format = '0.0000'  # current price
        if price_change is not None:
            ws.cell(row=row, column=5).number_format = '+0.0000;-0.0000'  # price change
        ws.cell(row=row, column=6).number_format = '$#,##0.00'  # cost
        ws.cell(row=row, column=7).number_format = '0.00'       # shares
        ws.cell(row=row, column=8).number_format = '+$#,##0.000;-$#,##0.000'  # expected pnl
        if realized_pnl is not None:
            ws.cell(row=row, column=10).number_format = '+$#,##0.000;-$#,##0.000'  # realized pnl

    # Summary row
    stats = positions_data["stats"]
    summary_row = len(positions) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=12)
    ws.cell(row=summary_row + 1, column=1, value="Balance:")
    ws.cell(row=summary_row + 1, column=2, value=f"${stats['current_balance']:.2f}")
    ws.cell(row=summary_row + 2, column=1, value="Invested:")
    ws.cell(row=summary_row + 2, column=2, value=f"${total_cost:.2f}")
    ws.cell(row=summary_row + 3, column=1, value="Expected PnL (if all win):")
    ws.cell(row=summary_row + 3, column=2, value=f"${total_expected_pnl:.2f}")
    ws.cell(row=summary_row + 4, column=1, value="Realized PnL (resolved):")
    ws.cell(row=summary_row + 4, column=2, value=f"${total_realized_pnl:.2f}")
    ws.cell(row=summary_row + 5, column=1, value="Wins / Losses:")
    ws.cell(row=summary_row + 5, column=2, value=f"{stats['wins']} / {stats['losses']}")
    ws.cell(row=summary_row + 6, column=1, value="Updated:")
    ws.cell(row=summary_row + 6, column=2, value=datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"))

    for r in range(summary_row, summary_row + 7):
        ws.cell(row=r, column=1).font = Font(bold=True)

    # Column widths
    widths = [4, 50, 11, 11, 12, 10, 8, 12, 10, 12, 10, 14, 18, 9, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64 + i) if i <= 26 else ""].width = w
    # Fix column letters for all
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze top row
    ws.freeze_panes = "A2"

    # Auto-filter
    ws.auto_filter.ref = f"A1:O{len(positions) + 1}"

    wb.save(EXCEL_FILE)
    log.info("Saved: %s (%d positions)", EXCEL_FILE, len(positions))


if __name__ == "__main__":
    data = load_positions()
    build_excel(data)
